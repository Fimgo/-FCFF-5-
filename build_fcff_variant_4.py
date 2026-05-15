from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT_DIR = Path(__file__).resolve().parent
OUT_FILE = OUT_DIR / "FCFF_варіант_4_готово.xlsx"


def round1(value):
    return round(value, 1)


def build_forecast():
    assumptions = {
        "revenue": 1120.0,
        "opex": 699.0,
        "depr": 70.0,
        "capex": 103.0,
        "delta_wc": 32.0,
        "revenue_growth": 0.058,
        "capex_pct_revenue": 0.087,
        "delta_wc_pct_revenue_growth": 0.024,
        "tax_rate": 0.18,
    }

    opex_pct = assumptions["opex"] / assumptions["revenue"]
    depr_pct = assumptions["depr"] / assumptions["revenue"]

    rows = []
    base_ebit = assumptions["revenue"] - assumptions["opex"] - assumptions["depr"]
    base_tax = base_ebit * assumptions["tax_rate"]
    base_nopat = base_ebit - base_tax
    base_fcff = (
        base_nopat
        + assumptions["depr"]
        - assumptions["capex"]
        - assumptions["delta_wc"]
    )

    rows.append(
        {
            "period": "Базовий рік",
            "revenue": assumptions["revenue"],
            "opex": assumptions["opex"],
            "ebit": base_ebit,
            "tax": base_tax,
            "nopat": base_nopat,
            "depr": assumptions["depr"],
            "capex": assumptions["capex"],
            "delta_wc": assumptions["delta_wc"],
            "fcff": base_fcff,
        }
    )

    previous_revenue = assumptions["revenue"]
    for year in range(1, 6):
        revenue = previous_revenue * (1 + assumptions["revenue_growth"])
        opex = revenue * opex_pct
        depr = revenue * depr_pct
        ebit = revenue - opex - depr
        tax = ebit * assumptions["tax_rate"]
        nopat = ebit - tax
        capex = revenue * assumptions["capex_pct_revenue"]
        delta_wc = (revenue - previous_revenue) * assumptions["delta_wc_pct_revenue_growth"]
        fcff = nopat + depr - capex - delta_wc

        rows.append(
            {
                "period": f"Рік {year}",
                "revenue": revenue,
                "opex": opex,
                "ebit": ebit,
                "tax": tax,
                "nopat": nopat,
                "depr": depr,
                "capex": capex,
                "delta_wc": delta_wc,
                "fcff": fcff,
            }
        )
        previous_revenue = revenue

    assumptions["opex_pct_revenue"] = opex_pct
    assumptions["depr_pct_revenue"] = depr_pct
    return assumptions, rows


def apply_table_style(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")


def main():
    assumptions, rows = build_forecast()
    wb = Workbook()
    ws = wb.active
    ws.title = "FCFF Model"
    ws_assumptions = wb.create_sheet("Assumptions")
    ws_method = wb.create_sheet("Method")

    dark_blue = "1F4E78"
    light_green = "E2F0D9"
    light_yellow = "FFF2CC"
    light_gray = "F2F2F2"

    title_font = Font(size=14, bold=True, color=dark_blue)
    white_bold = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    input_font = Font(color="0000FF")

    # Assumptions sheet
    ws_assumptions.sheet_view.showGridLines = False
    ws_assumptions["A1"] = "Вхідні дані та прогнозні припущення: варіант 4"
    ws_assumptions["A1"].font = title_font
    ws_assumptions.merge_cells("A1:D1")

    assumption_rows = [
        ("Показник", "Значення", "Одиниця", "Джерело/логіка"),
        ("Виручка", assumptions["revenue"], "млн грн", "Початкові дані варіанта 4"),
        ("Операційні витрати", assumptions["opex"], "млн грн", "Початкові дані варіанта 4"),
        ("Амортизація", assumptions["depr"], "млн грн", "Початкові дані варіанта 4"),
        ("CapEx", assumptions["capex"], "млн грн", "Початкові дані варіанта 4"),
        ("ΔWC", assumptions["delta_wc"], "млн грн", "Початкові дані варіанта 4"),
        ("Темп росту виручки", assumptions["revenue_growth"], "%", "Індивідуальне припущення"),
        ("CapEx як % від виручки", assumptions["capex_pct_revenue"], "%", "Індивідуальне припущення"),
        (
            "ΔWC як % від приросту виручки",
            assumptions["delta_wc_pct_revenue_growth"],
            "%",
            "Індивідуальне припущення",
        ),
        ("Податкова ставка", assumptions["tax_rate"], "%", "Індивідуальне припущення"),
        ("Операційні витрати як % виручки", assumptions["opex_pct_revenue"], "%", "Розраховано з базового року"),
        ("Амортизація як % виручки", assumptions["depr_pct_revenue"], "%", "Розраховано з базового року"),
    ]
    for r, row in enumerate(assumption_rows, start=3):
        for c, value in enumerate(row, start=1):
            ws_assumptions.cell(r, c, value)

    for cell in ws_assumptions[3]:
        cell.fill = PatternFill("solid", fgColor=dark_blue)
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    apply_table_style(ws_assumptions, 3, 15, 1, 4)
    for row in range(4, 13):
        ws_assumptions[f"B{row}"].font = input_font
        ws_assumptions[f"A{row}"].fill = PatternFill("solid", fgColor=light_yellow)
        ws_assumptions[f"B{row}"].fill = PatternFill("solid", fgColor=light_yellow)
    for row in [10, 11, 12, 13, 14, 15]:
        ws_assumptions[f"B{row}"].number_format = "0.0%"
    for row in range(4, 10):
        ws_assumptions[f"B{row}"].number_format = "#,##0.0"
    ws_assumptions.column_dimensions["A"].width = 34
    ws_assumptions.column_dimensions["B"].width = 16
    ws_assumptions.column_dimensions["C"].width = 12
    ws_assumptions.column_dimensions["D"].width = 34

    # Main model sheet: numeric output only, no fragile Excel formulas.
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Модель FCFF: п'ятирічний прогноз"
    ws["A1"].font = title_font

    headers = ["Показник", "Початкове значення", "Рік 1", "Рік 2", "Рік 3", "Рік 4", "Рік 5"]
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col, header)

    metric_map = [
        ("Рік", "period"),
        ("Виручка", "revenue"),
        ("Операційні витрати", "opex"),
        ("EBIT", "ebit"),
        ("Податки", "tax"),
        ("EBIT (1 - Tax)", "nopat"),
        ("Амортизація", "depr"),
        ("CapEx", "capex"),
        ("ΔWC", "delta_wc"),
        ("FCFF", "fcff"),
    ]
    for r_idx, (label, key) in enumerate(metric_map, start=5):
        ws.cell(r_idx, 1, label)
        for c_idx, period_data in enumerate(rows, start=2):
            value = period_data[key]
            ws.cell(r_idx, c_idx, value if key == "period" else round1(value))

    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=dark_blue)
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    apply_table_style(ws, 4, 14, 1, 7)
    for row in range(6, 15):
        for col in range(2, 8):
            ws.cell(row, col).number_format = "#,##0.0"
            ws.cell(row, col).alignment = Alignment(horizontal="right", vertical="center")
    for col in range(2, 8):
        ws.cell(5, col).alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[14]:
        cell.fill = PatternFill("solid", fgColor=light_green)
        cell.font = bold

    ws["A17"] = "Графік FCFF"
    ws["A17"].font = Font(size=12, bold=True, color=dark_blue)
    chart = LineChart()
    chart.title = "Прогноз FCFF, млн грн"
    chart.y_axis.title = "млн грн"
    chart.x_axis.title = "Рік"
    data = Reference(ws, min_col=3, max_col=7, min_row=14, max_row=14)
    categories = Reference(ws, min_col=3, max_col=7, min_row=5, max_row=5)
    chart.add_data(data, from_rows=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 17
    chart.series[0].graphicalProperties.line.solidFill = dark_blue
    chart.series[0].graphicalProperties.line.width = 25000
    chart.series[0].marker.symbol = "circle"
    chart.series[0].marker.size = 6
    ws.add_chart(chart, "A18")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "C5"

    # Method sheet with clear formulas written as text for audit.
    ws_method.sheet_view.showGridLines = False
    ws_method["A1"] = "Методика розрахунку"
    ws_method["A1"].font = title_font
    ws_method.merge_cells("A1:C1")
    method_rows = [
        ("Показник", "Формула", "Пояснення"),
        ("Виручка t", "Виручка t-1 × (1 + темп росту виручки)", "Темп росту = 5.8%"),
        ("Операційні витрати t", "Виручка t × (операційні витрати базового року / виручка базового року)", "Частка витрат стала"),
        ("Амортизація t", "Виручка t × (амортизація базового року / виручка базового року)", "Частка амортизації стала"),
        ("EBIT", "Виручка - операційні витрати - амортизація", "Операційний прибуток"),
        ("Податки", "EBIT × 18%", "Податкова ставка варіанта 4"),
        ("EBIT (1 - Tax)", "EBIT - податки", "NOPAT"),
        ("CapEx t", "Виручка t × 8.7%", "CapEx як % від виручки"),
        ("ΔWC t", "(Виручка t - виручка t-1) × 2.4%", "Зміна робочого капіталу"),
        ("FCFF", "EBIT (1 - Tax) + амортизація - CapEx - ΔWC", "Вільний грошовий потік для фірми"),
    ]
    for r, row in enumerate(method_rows, start=3):
        for c, value in enumerate(row, start=1):
            ws_method.cell(r, c, value)
    for cell in ws_method[3]:
        cell.fill = PatternFill("solid", fgColor=dark_blue)
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    apply_table_style(ws_method, 3, 12, 1, 3)
    for row in range(4, 13):
        ws_method[f"A{row}"].fill = PatternFill("solid", fgColor=light_gray)
        ws_method[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_method[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_method.row_dimensions[row].height = 34
    ws_method.column_dimensions["A"].width = 24
    ws_method.column_dimensions["B"].width = 58
    ws_method.column_dimensions["C"].width = 32

    for sheet in wb.worksheets:
        for row in range(1, sheet.max_row + 1):
            if sheet.row_dimensions[row].height is None:
                sheet.row_dimensions[row].height = 21

    wb.save(OUT_FILE)
    print("ok")


if __name__ == "__main__":
    main()
